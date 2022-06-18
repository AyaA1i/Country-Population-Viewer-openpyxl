# This program shows the population of states in different countries, by loading these information from an excel file
# Authors: Youssef Morad, Aya Ali, Mariam Ayman
# Date: 05/03/2022

from openpyxl import load_workbook
from os import system, name

wb = load_workbook('file.xlsx')
ls = []
for name in wb.sheetnames:
    name = name.lower()
    ls.append(name)


def exit_or_not():
    print("\n3: to choose another country")
    print("4: to exit")
    ex = input('Please choose either 3 or 4: ')
    while ex != '3' and ex != '4':
        ex = input('Please choose either 3 or 4: ')

    return ex


def clear():
    if name == 'nt':
        _ = system('cls')


def num_of_rows(country, ws):
    rows = 0
    cell = ws['A1'].value
    while cell != None:
        rows += 1
        cell = ws['A' + str(rows + 1)].value
    return rows


def option_1(country_name, ws):
    countries_pop = '\n'
    Total_Population = 0
    rows = num_of_rows(country_name, ws)
    for row in range(2, rows):
        cell1 = ws['A' + str(row)].value
        cell2 = ws['B' + str(row)].value
        countries_pop += ("{0:20} {1:2}\n".format(cell1, cell2))
    for row in range(2, rows):
        cell = ws['B' + str(row)].value
        Total_Population += cell
    total_pop_text = f'\nTotal Population of {country_name.title()} is "{Total_Population}"'
    return (countries_pop + total_pop_text)


def option_2(country, ws):
    rows = num_of_rows(country, ws)
    max = 0
    min = ws['B2'].value
    for row in range(2, rows):
        cell = ws['B' + str(row)].value
        if cell > max:
            max = cell
            max_row = row
        if cell < min:
            min = cell
            min_row = row
    min_country_name = ws['A' + str(min_row)].value
    max_country_name = ws['A' + str(max_row)].value
    lh_pop = f'''\nThe lowest population in {country.title()} is "{min}" in "{min_country_name}" \n'''
    lh_pop += f'''\nThe highest population in {country.title()} is "{max}" in "{max_country_name}" \n '''
    return lh_pop


def excel_program():
    while True:
        country = input("Please choose a country to load its file: ").lower().strip()
        while country not in ls:
            print("The chosen country isn't available")
            country = input("Please choose another country: ").lower().strip()
        ws = wb[country.title()]
        print("\n1: To display the population of each province and the total population of the country")
        print("2: To display the province with the highest and lowest population")
        option = input("Please choose either 1 or 2: ")
        while option != '1' and option != '2':
            option = input("\nPlease choose either 1 or 2: ")
        if option == '1':
            clear()
            x = option_1(country, ws)
            print(x)
            ex = exit_or_not()
            if ex == '3':
                print()
                clear()
            else:
                break
        elif option == '2':
            clear()
            print()
            y = option_2(country, ws)
            print(y)
            ex = exit_or_not()
            if ex == '3':
                print()
                clear()
            else:
                break


